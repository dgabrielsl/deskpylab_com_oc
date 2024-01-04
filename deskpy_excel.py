import os
import sqlite3

from PyQt6.QtWidgets import QFileDialog, QMessageBox

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font

import re

class Excel():
    def load_sysde(self):
        con = sqlite3.connect('sysde.db')
        cur = con.cursor()

        try:
            cur.execute('''
                CREATE TABLE sysde_hub(
                    IDENT VARCHAR(25) UNIQUE,
                    EMAIL VARCHAR(50),
                    PHONE VARCHAR(25))
                ''')
        except: pass

        path = QFileDialog.getOpenFileName(filter=('*.xlsx'))
        path = path[0]

        self.statusbar.showMessage(f'Loading new Sysde workbook: «{path}»',9000)

        wb = openpyxl.load_workbook(path)
        ws = wb.worksheets[0]

        ws.delete_rows(1,4)

        saved_copy = path
        saved_copy = saved_copy.split('/')
        saved_copy.pop()
        saved_copy = '/'.join(saved_copy)
        saved_copy = f'{saved_copy}/temp_copy.xlsx'

        wb.save(saved_copy)

        for i in range(ws.max_column):
            i += 1
            if ws.cell(1,i).value == 'Identificación' or ws.cell(1,i).value == 'Identificacion': char_1 = ws.cell(1,i).column_letter
            if ws.cell(1,i).value == 'Email': char_2 = ws.cell(1,i).column_letter
            if ws.cell(1,i).value == 'Teléfono celular' or ws.cell(1,i).value == 'Telefono celular': char_3 = ws.cell(1,i).column_letter

        self.records = []

        for i in range(int(ws.max_row) + 1):
            if i > 1:
                line = []

                # Identification:
                insert = f'{ws[char_1+str(i)].value}'
                insert = insert.replace('-','')
                line.append(str(insert))

                # E-mail:
                line.append(f'{ws[char_2+str(i)].value}'.lower())

                # Phone:
                line.append(str(f'{ws[char_3+str(i)].value}'))

                self.records.append(line)

        for record in self.records:
            r = f'INSERT INTO sysde_hub VALUES ("{record[0]}", "{record[1]}", "{record[2]}")'
            try: cur.execute(r)
            except Exception as e: pass

        con.commit()
        con.close()

        try: os.remove(saved_copy)
        except Exception as e: pass

        QMessageBox.information(
            self,
            'deskpy_excel',
            'La actualización de la base de datos de SYSDE se ha completado correctamente.\t\t',
            QMessageBox.StandardButton.Ok,
            QMessageBox.StandardButton.Ok)

    def load_book(self):
        wb_url = QFileDialog.getOpenFileName(filter=('*.xlsx'))
        wb_url = wb_url[0]

        con = sqlite3.connect('hub.db')
        cur = con.cursor()

        try:
            cur.execute('''
                CREATE TABLE customers(
                    LOAD_IDENTIFIER VARCHAR(100),
                    HELPDESK VARCHAR(10) UNIQUE,
                    IDENTIFICATION VARCHAR(20),
                    DOCUMENT VARCHAR(10),
                    CODE VARCHAR(10),
                    CLASS_CASE VARCHAR(100),
                    STATUS VARCHAR(20),
                    PRODUCT VARCHAR(20),
                    INCOME_SOURCE VARCHAR(300),
                    WARNING_AMOUNT VARCHAR(20),
                    CUSTOMER_PROFILE VARCHAR(200),
                    DEADLINE VARCHAR(20),
                    NOTIF_TYPE VARCHAR(20),
                    CONTACT_TYPE VARCHAR(20),
                    CUSTOMER_ANSWER VARCHAR(200),
                    ASSIGNED_TO VARCHAR(50),
                    AUTHOR VARCHAR(50),
                    RESULT VARCHAR(100),
                    UPDATED VARCHAR(20),
                    FNAME VARCHAR(150),
                    CHANGES_LOG VARCHAR(5000))
                ''')
        except Exception as e: pass

        try:
            wb = openpyxl.load_workbook(wb_url)
            ws = wb.worksheets[0]
        except Exception as e: print(e)

        helpdesk = ''
        identification = ''
        document = ''
        code = ''
        class_case = ''
        status = ''
        product = ''
        income_source = ''
        warning_amount = ''
        customer_profile = ''
        deadline = ''
        notif_type = ''
        contact_type = ''
        customer_answer = ''
        assigned_to = ''
        author = ''
        result = ''
        updated = ''
        fname = ''

        for i in range(ws.max_column):
            i += 1
            value = ws.cell(1,i).value.lower()
            value = value.replace(':','').replace('.','')
            value = value.replace('á','a').replace('é','e').replace('í','i').replace('ó','o').replace('ú','u')

            if value.__contains__('#'): helpdesk = ws.cell(1,i).column_letter
            if value.__contains__('cedula'): identification = ws.cell(1,i).column_letter
            if value.__contains__('pagare'): document = ws.cell(1,i).column_letter
            if value.__contains__('codigo de cliente'): code = ws.cell(1,i).column_letter
            if value.__contains__('tipo de caso'): class_case = ws.cell(1,i).column_letter
            if value.__contains__('estado'): status = ws.cell(1,i).column_letter
            if value.__contains__('producto'): product = ws.cell(1,i).column_letter
            if value.__contains__('origen de fondos'): income_source = ws.cell(1,i).column_letter
            if value.__contains__('monto de la alerta'): warning_amount = ws.cell(1,i).column_letter
            if value.__contains__('perfil del cliente'): customer_profile = ws.cell(1,i).column_letter
            if value.__contains__('fecha de prorroga'): deadline = ws.cell(1,i).column_letter
            if value.__contains__('tipo de notificacion'): notif_type = ws.cell(1,i).column_letter
            if value.__contains__('tipo de contacto'): contact_type = ws.cell(1,i).column_letter
            if value.__contains__('respuesta del cliente'): customer_answer = ws.cell(1,i).column_letter
            if value.__contains__('asignado a'): assigned_to = ws.cell(1,i).column_letter
            if value.__contains__('autor'): author = ws.cell(1,i).column_letter
            if value.__contains__('resultado de gestion'): result = ws.cell(1,i).column_letter
            if value.__contains__('actualizado'): updated = ws.cell(1,i).column_letter
            if value.__contains__('asunto'): fname = ws.cell(1,i).column_letter

        self.customers = []

        def reduce_blanks(raw_string):
            strip_str = raw_string

            if strip_str.__contains__('  '):
                raw_string = []
                strip_str = strip_str.split(' ')

                for ss in strip_str:
                    if ss == '' or ss == '\t' or ss == '\n' or ss == '\r' or ss == '\f' or ss == '\v' or len(ss) < 1: pass
                    else: raw_string.append(ss)

                strip_str.clear()
                raw_string = ' '.join(raw_string)

        for i in range(int(ws.max_row) + 1):
            if i > 1:
                line = []

                # HelpDesk / Don't clear.
                insert = f'{ws[helpdesk+str(i)].value}'
                if insert == '' or insert == None or insert == 'None' or insert == 'NONE': insert = ''
                # print(f'hd→{insert}')
                line.append(insert)

                # Identification / Clear: \s - . ,
                insert = f'{ws[identification+str(i)].value}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else: insert = insert.strip().replace('-','').replace('.','').replace(',','')

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'identification→{insert}')
                line.append(insert)

                # Document / Clear: \s / N n A a
                insert = f'{ws[document+str(i)].value}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else: insert = insert.strip().replace('N','').replace('n','').replace('A','').replace('a','').replace('/','')

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                if not re.search(r'\d', insert): insert = ''
                if insert == '0' or insert == 0: insert = ''
                # print(f'document→{insert}')
                line.append(insert)

                # Code / Clear: \s / N n A a
                insert = f'{ws[code+str(i)].value}'
                
                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else:
                    insert = insert.strip().replace('N','').replace('n','').replace('A','').replace('a','').replace('/','')
                    if insert == 0 or insert == '0' or insert == 'None' or insert == 'NONE': insert = ''

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'code→{insert}')
                line.append(insert)

                # Class case.
                insert = f'{ws[class_case+str(i)].value}'
                if insert == None or insert == 'None' or insert == 'NONE': insert = ''

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                if insert == '0' or insert == 0: insert = ''
                # print(f'class case→{insert}')
                line.append(insert)

                # Status / Prevent: \s and Customize: to uppercase
                insert = f'{ws[status+str(i)].value}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else: insert = insert.strip().upper()

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'status→{insert}')
                line.append(insert)

                # Product / Prevent: \s and Customize: to uppercase
                insert = f'{ws[product+str(i)].value}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else: insert = insert.strip().upper()

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'product→{insert}')
                line.append(insert)

                # Income source / Customize: to uppercase
                insert = f'{ws[income_source+str(i)].value}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else: insert = insert.replace('N/A','').replace('n/a','').replace('N/a','').replace('n/A','').upper()

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'income source→{insert}')
                line.append(insert)

                # Warning amount / Clear: \s N n A a ? ¢ / $
                # Search pattern with decimal's amounts to remove it.
                insert = f'{ws[warning_amount+str(i)].value}'

                if insert.lower().__contains__('alert') or insert.lower().__contains__('dupl'):
                    insert = 'ALERTA DUPLICADA'
                    line.append(insert)
                else:
                    # Check if there's any ¢ or $ special character to add at the end:
                    sfx = ''
                    if insert.__contains__('¢'): sfx = 'CRC'
                    elif insert.__contains__('$'): sfx = 'USD'

                    # Normalice all to dots:
                    insert = insert.replace(',','.')

                    # Remove any character if isn't digit:
                    insert = insert.replace('/','').replace('¢','').replace('$','').replace('?','')
                    insert = insert.replace('N','').replace('n','').replace('A','').replace('a','')

                    # Build up the patterns to avoid float:
                    match_a_dot = re.search(r'\.\d$', insert)
                    match_b_dot = re.search(r'\.\d\d$', insert)
                    match_a_spc = re.search(r'\s\d$', insert)
                    match_b_spc = re.search(r'\s\d\d$', insert)

                    # Removing decimals:
                    if match_a_dot or match_a_spc: insert = insert[:-2]
                    elif match_b_dot or match_b_spc: insert = insert[:-3]

                    # Full cleaning keeping just digits:
                    insert = insert.replace(' ','').replace('.','')

                    # Split miles by dots:
                    if len(insert) == 4: insert = f'{insert[0]}.{insert[1:]}'                           # 1.000
                    elif len(insert) == 5: insert = f'{insert[:2]}.{insert[2:]}'                        # 10.000
                    elif len(insert) == 6: insert = f'{insert[:3]}.{insert[3:]}'                        # 100.000
                    elif len(insert) == 7: insert = f'{insert[0]}.{insert[1:4]}.{insert[4:]}'           # 1.000.000
                    elif len(insert) == 8: insert = f'{insert[:2]}.{insert[2:5]}.{insert[5:]}'          # 10.000.000
                    elif len(insert) == 9: insert = f'{insert[:3]}.{insert[3:6]}.{insert[6:]}'          # 100.000.000

                    # More filters:
                    v = ws[warning_amount+str(i)].value
                    if v == None or str(v) == '0' or insert == 'None' or insert == 'NONE': insert = ''

                    strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                    if strip_insert == '': insert = ''
                    reduce_blanks(insert)
                    # print(f'warning amount→{insert}')
                    if sfx != '': line.append(f'{insert} {sfx}')
                    else: line.append(insert)

                # Customer profile.
                insert = f'{ws[customer_profile+str(i)].value}'
                if insert == None or insert == 'None' or insert == 'NONE': insert = ''

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'customer profile→{insert}')
                line.append(insert)

                # Deadline / Fix: save as dd/mm/yyyy
                insert = f'{ws[deadline+str(i)].value}'

                # if insert == None or insert == '' or insert == 'None' or insert == 'NONE': insert = ''
                # else:
                #     if re.search(r'^(\d{1,2}\/\d{1,2}\/\d{1,4})', insert):
                #         insert = insert.split('/')
                #         insert = f'{insert[1]}/{insert[0]}/{insert[2]}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else:
                    insert = insert.split(' ')
                    insert = insert[0]

                    if insert.__contains__('/'): insert = insert.split('/')
                    elif insert.__contains__('-'): insert = insert.split('-')

                    try: insert = f'{insert[2]}/{insert[1]}/{insert[0]}'
                    except: pass

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'deadline→{insert}')
                line.append(insert)

                # Notification type / Clear: N n A a /
                insert = f'{ws[notif_type+str(i)].value}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else: insert = insert.replace(' ','').replace('N','').replace('n','').replace('A','').replace('a','').replace('/','')

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'notification type→{insert}')
                line.append(insert)

                # Contact type / Clear: N n A a /               
                insert = f'{ws[contact_type+str(i)].value}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else: insert = insert.replace(' ','').replace('N','').replace('n','').replace('A','').replace('a','').replace('/','')

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'contact type→{insert}')
                line.append(insert)

                # Customer answer / Clear: prefix number, dot and sometimes \s at the beggining of the text by pattern searching.
                insert = f'{ws[customer_answer+str(i)].value}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else:
                    if re.search(r'^\d\.\s', insert): insert = insert[3:]
                    elif re.search(r'^\d\.\D', insert): insert = insert[2:]

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'customer answer→{insert}')
                line.append(insert)

                # Assigned to / Normalize to lowercase, then capitalize
                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else:
                    insert = f'{ws[assigned_to+str(i)].value}'
                    insert = insert.lower()
                    insert = insert.title()

                strip_insert = insert.replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'assigned to→{insert}')
                line.append(insert)

                # Author / Normalize to lowercase, then title
                insert = f'{ws[author+str(i)].value}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else:
                    insert = insert.lower()
                    insert = insert.title()

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'author→{insert}')
                line.append(insert)

                # Result / Clear: prefix number, dot and sometimes \s at the beggining of the text by pattern searching.
                insert = f'{ws[result+str(i)].value}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else:
                    if re.search(r'^\d\.\s', insert): insert = insert[3:]
                    elif re.search(r'^\d\.\D', insert): insert = insert[2:]

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'result→{insert}')
                line.append(insert)

                # Updated / Clean: time, keep just date; Fix: save as dd/mm/yyyy
                insert = f'{ws[updated+str(i)].value}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else:
                    insert = insert.split(' ')
                    insert = insert[0]

                    if insert.__contains__('/'): insert = insert.split('/')
                    elif insert.__contains__('-'): insert = insert.split('-')

                    try: insert = f'{insert[2]}/{insert[1]}/{insert[0]}'
                    except: pass

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')
                if strip_insert == '': insert = ''
                # print(f'updated→{insert}')
                line.append(insert)

                # Full name (subject) / Clean: if not name/lastname
                insert = f'{ws[fname+str(i)].value}'

                if insert == None or insert == 'None' or insert == 'NONE': insert = ''
                else:
                    rem_s_insert = insert.split(' ')
                    insert = []
                    for rsi in rem_s_insert:
                        if len(rsi) > 0: insert.append(rsi)

                    insert = ' '.join(insert)
                    insert = insert.upper()

                strip_insert = insert.replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')

                if strip_insert == '': insert = ''
                else:
                    try:
                        insert = insert.lower()
                        for c in self.dict_raw_txt:
                            insert = insert.replace(c,'')
                    except Exception as e: print(e)
                    insert = insert.upper()

                # print(f'fname→{insert}')
                line.append(insert)

                depured_line = []

                for l in line:
                    _l_ = l.strip().replace(' ','').replace('\n','').replace('\t','').replace('\r','').replace('\f','').replace('\v','')

                    reduce_blanks(l)

                    if _l_ == '': depured_line.append('')
                    else:
                        strip_str = l

                        if strip_str.__contains__('  ') or strip_str.__contains__('\t') or strip_str.__contains__('\n') or strip_str.__contains__('\r') or strip_str.__contains__('\f') or strip_str.__contains__('\v'):
                            l = []
                            strip_str = strip_str.split(' ')

                            for ss in strip_str:
                                if ss.__contains__('') or ss.__contains__('\t') or ss.__contains__('\n') or ss.__contains__('\r') or ss.__contains__('\f') or ss.__contains__('\v'): pass
                                else: l.append(ss.replace(' ','').replace('\t',''))

                            strip_str.clear()
                            l = ' '.join(l)

                        depured_line.append(l)

                line.clear()
                line = depured_line
                depured_line = []

                # print(f'>>> HD:\t\t{line[0]}')
                self.customers.append(line)

        self.logs_count.setText(str(len(self.customers)))

        QMessageBox.information(
            self,
            'DeskPyL COM',
            f'\nLa carga ha sido completada.\t\n{len(self.customers)} nuevos registros preparados para guardar.\t\n',
            QMessageBox.StandardButton.Ok,
            QMessageBox.StandardButton.Ok)

        con.commit()
        con.close()

    def write_customers(self):
        query = self.load_tag_name.text()

        if not query.strip() == '':
            if len(query) > 9:
                con = sqlite3.connect('hub.db')
                cur = con.cursor()
                cur.execute('SELECT * FROM customers WHERE load_identifier = ?', (query,))
                res = cur.fetchone()

                if res == None:
                    if not re.search(r' ', query) and not re.search(r'\t', query):
                        for c in self.customers:
                            try:
                                record = f'INSERT INTO customers VALUES ("{query}", "{c[0]}", "{c[1]}", "{c[2]}", "{c[3]}", "{c[4]}", "{c[5]}", "{c[6]}", "{c[7]}", "{c[8]}", "{c[9]}", "{c[10]}", "{c[11]}", "{c[12]}", "{c[13]}", "{c[14]}", "{c[15]}", "{c[16]}", "{c[17]}", "{c[18]}", "")'
                                cur.execute(record)
                            except Exception as e: print(e)

                        QMessageBox.information(
                            self,
                            'DeskPyL COM',
                            f'\nRegistros procesados: {len(self.customers)}.\t\nNombre de la etiqueta: {query}.\t\n',
                            QMessageBox.StandardButton.Ok,
                            QMessageBox.StandardButton.Ok)
                        
                        self.load_tag_name.setText('')
                        self.logs_count.setText('')

                    else:
                        QMessageBox.information(
                            self,
                            'DeskPyL COM',
                            f'\nEl nombre de la etiqueta no puede contener espacios en blanco, tabulaciones o saltos de línea.\t\nPuede utilizar "_" en lugar de espacio.\t\n',
                            QMessageBox.StandardButton.Ok,
                            QMessageBox.StandardButton.Ok)
                else:
                    QMessageBox.information(
                        self,
                        'DeskPyL COM',
                        f'\nPara administrar más eficientemente los registrsos, el nombre de la etiqueta de carga debe ser única.\t\n"{query}" ya existe, por favor ingrese una etiqueta diferente para continuar.\t\n',
                        QMessageBox.StandardButton.Ok,
                        QMessageBox.StandardButton.Ok)

                con.commit()
                con.close()

            else:
                QMessageBox.information(
                    self,
                    'DeskPyL COM',
                    f'\nEl nombre de la etiqueta debe contener al menos 10 caracteres (letras).\t\n',
                    QMessageBox.StandardButton.Ok,
                    QMessageBox.StandardButton.Ok)
        else:
            QMessageBox.information(
                self,
                'DeskPyL COM',
                f'\nPor favor ingrese un nombre de etiqueta único para asociar a la carga de datos en proceso.\t\n',
                QMessageBox.StandardButton.Ok,
                QMessageBox.StandardButton.Ok)
        
